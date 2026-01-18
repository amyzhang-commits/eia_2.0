#!/usr/bin/env python3
"""
dataset_pipeline.py


Two-phase ingestion module for Excel/CSV datasets.

Phase 1: Schema Preview + Memory Estimation (NO merging)
         Column schemas, row counts, memory footprint, merge recommendations.
Phase 2: Merge files with coverage analysis and optional heatmap.

Designed for reproducibility, portability, and large public datasets (e.g. EIA).


Functions
---------
datasets_overview(folder_path=None)
    Inspect schemas + estimate memory usage without loading data.
convert_excel_to_csv(folder_path=None, output_folder=None, size_threshold_gb=0.5)
    Convert Excel files to CSV for memory-efficient chunked processing.
merge_folder_files(folder_path=None, save=False, chunk_size=None, show_preview=True)
    Memory-safe merge with summary of top 10 columns with most missing values (to guide first quick, obvious clean-up)
"""

from pathlib import Path
from typing import Optional, Union
import csv
import pandas as pd
from tqdm import tqdm
import openpyxl
import psutil
import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# ======================================================
# 📁 Project-relative paths
# ======================================================

FALLBACK_DATASETS_PATH = Path(__file__).parent.parent / "datasets_thermoelectric"
DATASETS_OUTPUT = Path(__file__).parent.parent / "datasets_merged"
DATASETS_OUTPUT.mkdir(exist_ok=True)


# ======================================================
# 🧠 Header detection logic
# ======================================================

def detect_header_row(
    file_path: Union[str, Path],
    nrows: int = 10,
    string_ratio_threshold: float = 0.5
) -> int:
    """Detect the most likely header row in a CSV or Excel file."""
    file_path = Path(file_path)

    try:
        if file_path.suffix.lower() == ".xlsx":
            preview = pd.read_excel(file_path, header=None, nrows=nrows)
        elif file_path.suffix.lower() == ".csv":
            preview = pd.read_csv(
                file_path,
                header=None,
                nrows=nrows,
                encoding="utf-8",
                errors="replace",
            )
        else:
            raise ValueError(f"Unsupported file type: {file_path.name}")

        for i, row in preview.iterrows():
            string_ratio = row.apply(lambda x: isinstance(x, str)).mean()
            if string_ratio > string_ratio_threshold:
                return i
        return 0
    except Exception as e:
        print(f"⚠️  Header detection failed for {file_path.name}: {e}")
        return 0


# ======================================================
# 📄 File row counters
# ======================================================

def get_total_rows_excel(file_path: Union[str, Path], header_row: int = 0) -> int:
    """Get exact data row count from Excel file (excluding header)."""
    file_path = Path(file_path)
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet = wb.active
        total_rows = sheet.max_row - header_row
        wb.close()
        return total_rows
    except Exception as e:
        print(f"⚠️  Row counting failed for {file_path.name} (Excel): {e}")
        return 0


def get_total_rows_csv(file_path: Union[str, Path], header_row: int = 0) -> int:
    """Get exact data row count from CSV file (excluding header)."""
    file_path = Path(file_path)
    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as fin:
            total_lines = sum(1 for _ in fin)
        return max(total_lines - (header_row + 1), 0)
    except Exception as e:
        print(f"⚠️  Row counting failed for {file_path.name} (CSV): {e}")
        return 0


# ======================================================
# 📄 Unified file reader
# ======================================================

def read_file(
    file_path: Union[str, Path],
    chunk_size: Optional[int] = None,
    verbose: bool = True
) -> tuple[pd.DataFrame, Optional[str]]:
    """Read a single file with auto-detected header.

    Returns
    -------
    tuple[pd.DataFrame, Optional[str]]
        (DataFrame, None) on success, (empty DataFrame, error message) on failure.
    """
    file_path = Path(file_path)
    header_row = detect_header_row(file_path)

    if verbose:
        print(f"📥 Reading {file_path.name} (header row: {header_row})")

    try:
        if file_path.suffix.lower() == ".xlsx":
            return pd.read_excel(file_path, header=header_row), None
        elif file_path.suffix.lower() == ".csv":
            if chunk_size:
                chunks = []
                for chunk in pd.read_csv(
                    file_path,
                    header=header_row,
                    encoding="utf-8",
                    errors="replace",
                    chunksize=chunk_size,
                ):
                    chunks.append(chunk)
                return pd.concat(chunks, ignore_index=True), None
            else:
                return pd.read_csv(
                    file_path,
                    header=header_row,
                    encoding="utf-8",
                    errors="replace",
                ), None
    except Exception as e:
        if verbose:
            print(f"⚠️  Read failed for {file_path.name}: {e}")
        return pd.DataFrame(), str(e)


# ======================================================
# 🔍 Phase 1: Schema Preview + Memory Estimation
# ======================================================

def datasets_overview(folder_path: Optional[Union[str, Path]] = None) -> None:
    """
    Inspect all files without merging. Provides schema analysis and memory estimation.

    Outputs:
    - Column counts and header detection per file
    - Schema consistency warnings
    - Memory footprint estimation
    - Merge safety recommendations

    Parameters
    ----------
    folder_path : str or pathlib.Path, optional
        Folder containing CSV/Excel files. If None, uses FALLBACK_DATASETS_PATH.
    """
    folder = Path(folder_path) if folder_path else FALLBACK_DATASETS_PATH
    files = [
        f
        for f in folder.glob("*")
        if f.suffix.lower() in [".xlsx", ".csv"] and not f.name.startswith(".")
    ]

    if not files:
        raise ValueError(f"No CSV or Excel files found in: {folder}")

    print("\n🔍 PHASE 1: DATASET OVERVIEW")
    print(f"   Target folder: {folder}")
    print("   This step inspects schemas and estimates memory usage only.")
    print("   No files are merged in Phase 1.\n")

    schemas = {}
    row_counts = {}
    mem_estimates_gb = {}
    large_excel_files = []

    for f in files:
        header_row = detect_header_row(f)

        try:
            # Header-only read to capture column names
            if f.suffix.lower() == ".xlsx":
                cols = pd.read_excel(f, header=header_row, nrows=0).columns.tolist()
                total_rows = get_total_rows_excel(f, header_row)
            else:
                cols = pd.read_csv(
                    f,
                    header=header_row,
                    nrows=0,
                    encoding="utf-8",
                    errors="replace",
                ).columns.tolist()
                total_rows = get_total_rows_csv(f, header_row)

            schemas[f.name] = cols
            row_counts[f.name] = total_rows

            # Memory estimate from 100-row sample
            if f.suffix.lower() == ".xlsx":
                sample = pd.read_excel(f, header=header_row, nrows=100)
            else:
                sample = pd.read_csv(
                    f,
                    header=header_row,
                    nrows=100,
                    encoding="utf-8",
                    errors="replace",
                )

            sample_bytes = sample.memory_usage(deep=True).sum()
            est_bytes = sample_bytes * (total_rows / 100) if total_rows > 0 else 0
            est_gb = est_bytes / 1e9
            mem_estimates_gb[f.name] = est_gb

            if f.suffix.lower() == ".xlsx" and est_gb > 0.5:
                large_excel_files.append((f.name, est_gb))

            print(
                f"   • {f.name}"
                f" | columns: {len(cols)}"
                f" | header row: {header_row}"
                f" | rows: {total_rows:,}"
                f" | estimated size: ~{est_gb:.2f} GB"
            )
        except Exception as e:
            print(f"⚠️  Skipping {f.name} due to error: {e}")
            continue

    # Schema consistency summary
    all_columns = set().union(*schemas.values())
    total_columns = len(all_columns)
    print(f"\n📊 SCHEMA SUMMARY")
    print(f"   Unique columns across all files: {total_columns}")

    for name, cols in schemas.items():
        missing = all_columns - set(cols)
        if missing:
            print(f"   ⚠️  {name} is missing {len(missing)} columns:")
            print(f"      {sorted(list(missing))}")

    # Memory recommendations
    total_est_gb = sum(mem_estimates_gb.values())
    system_ram_gb = psutil.virtual_memory().total / 1e9

    print("\n💾 MEMORY SUMMARY")
    print(f"   Estimated combined dataset size: ~{total_est_gb:.2f} GB")
    print(f"   Detected system RAM:            ~{system_ram_gb:.2f} GB")

    if large_excel_files:
        print("\n⚠️  EXCEL SIZE NOTICE")
        print("   The following Excel files are estimated > 0.5 GB each:")
        for fname, size in large_excel_files:
            print(f"      • {fname} (~{size:.2f} GB)")
        print("   Excel files cannot be read in chunks.")
        print("   Recommendation: Convert these to CSV before merging:")
        print("      convert_excel_to_csv(folder_path=..., output_folder=...)")

    if total_est_gb > 0.7 * system_ram_gb:
        suggested_chunk = max(
            1000, int(10000 * (system_ram_gb / total_est_gb))
        ) if total_est_gb > 0 else 10000
        print("\n⚠️  MEMORY RECOMMENDATION")
        print("   Direct, in-memory merge may be unsafe.")
        print(f"   Recommendation: Use a chunked merge with:")
        print(f"      merge_folder_files(chunk_size={suggested_chunk})")
        if large_excel_files:
            print("   (Run convert_excel_to_csv first for large Excel files.)")
    else:
        print("\n✅ MEMORY RECOMMENDATION")
        print("   Direct merge is likely safe for this dataset:")
        print("      merge_folder_files(chunk_size=None)")

    if total_columns > 30:
        print("\n🗄️  STORAGE RECOMMENDATION")
        print(f"   Detected a wide schema with {total_columns} columns.")
        print("   Consider persisting to a database or using a columnar store.")
        print("   Example: prepare for SQL ingestion with a dedicated utility.")

    print("\n✅ PHASE 1 COMPLETE")
    print("   Review the above summary before running the merge step.\n")


# ======================================================
# 🔄 Excel to CSV Converter
# ======================================================

def convert_excel_to_csv(
    folder_path: Optional[Union[str, Path]] = None,
    output_folder: Optional[Union[str, Path]] = None,
    size_threshold_gb: float = 0.5
) -> None:
    """
    Convert Excel files to CSV format for memory-efficient chunked processing.

    Useful when Excel files are too large for memory and need chunked reading.
    Converts files above the size threshold and saves them to output folder.

    Parameters
    ----------
    folder_path : str or pathlib.Path, optional
        Folder containing Excel files. If None, uses FALLBACK_DATASETS_PATH.
    output_folder : str or pathlib.Path, optional
        Where to save CSV files. If None, saves to same folder as Excel files.
    size_threshold_gb : float, default 0.5
        Minimum file size (GB) to trigger conversion. Files smaller than this are skipped.
    """
    folder = Path(folder_path) if folder_path else FALLBACK_DATASETS_PATH
    output = Path(output_folder) if output_folder else folder
    output.mkdir(exist_ok=True)

    excel_files = [f for f in folder.glob("*.xlsx") if not f.name.startswith(".")]

    if not excel_files:
        print(f"No Excel files found in {folder}")
        return

    print(f"\n🔄 EXCEL → CSV CONVERSION")
    print(f"   Input folder:  {folder}")
    print(f"   Output folder: {output}")
    print(f"   Size threshold: {size_threshold_gb} GB\n")

    converted_count = 0
    skipped_count = 0

    for f in excel_files:
        header_row = detect_header_row(f)
        total_rows = get_total_rows_excel(f, header_row)

        try:
            sample = pd.read_excel(f, header=header_row, nrows=100)
            sample_bytes = sample.memory_usage(deep=True).sum()
            est_bytes = sample_bytes * (total_rows / 100) if total_rows > 0 else 0
            est_gb = est_bytes / 1e9

            if est_gb < size_threshold_gb:
                print(
                    f"⏭️  Skipping {f.name} "
                    f"(~{est_gb:.2f} GB < {size_threshold_gb:.2f} GB threshold)"
                )
                skipped_count += 1
                continue

            print(f"🔄 Converting {f.name} (~{est_gb:.2f} GB)...")

            # Stream rows with openpyxl to avoid loading entire file into memory
            wb = openpyxl.load_workbook(f, read_only=True)
            sheet = wb.active
            output_file = output / f"{f.stem}.csv"

            with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                for i, row in enumerate(sheet.iter_rows(values_only=True)):
                    if i < header_row:
                        continue  # skip pre-header rows
                    writer.writerow(row)

            wb.close()

            print(f"   ✅ Saved CSV: {output_file.name}")
            converted_count += 1

        except Exception as e:
            print(f"   ⚠️  Conversion failed for {f.name}: {e}")
            continue

    print("\n✅ CONVERSION SUMMARY")
    print(f"   Converted: {converted_count} file(s)")
    print(f"   Skipped:   {skipped_count} file(s)")
    print(f"   Output:    {output}\n")


# ======================================================
# 🔗 Phase 2: Merge (with optional top 10 missing preview)
# ======================================================

def merge_folder_files(
    folder_path: Optional[Union[str, Path]] = None,
    save: bool = False,
    chunk_size: Optional[int] = None,
    show_preview: bool = True
) -> pd.DataFrame:
    """
    Merge all CSV/Excel files in a folder with optional chunked reading.
    Optionally prints top 10 missing counts + percentages.

    Parameters
    ----------
    folder_path : str or pathlib.Path, optional
        Folder containing files to merge. If None, uses FALLBACK_DATASETS_PATH.
    save : bool, default False
        Whether to save merged DataFrame to CSV.
    chunk_size : int, optional
        Number of rows to read at a time for CSV files.
    show_preview : bool, default True
        If True, prints top 10 missing counts + percentages.

    Returns
    -------
    pd.DataFrame
        Merged DataFrame containing all file data.
    """
    folder = Path(folder_path) if folder_path else FALLBACK_DATASETS_PATH
    files = [
        f
        for f in folder.glob("*")
        if f.suffix.lower() in [".xlsx", ".csv"] and not f.name.startswith(".")
    ]

    if not files:
        raise ValueError(f"No valid files found in {folder}")

    print(f"\n🔗 PHASE 2: MERGE")
    print(f"   Source folder: {folder}")
    if chunk_size:
        print(f"   Chunked reading enabled: {chunk_size:,} rows per chunk\n")
    else:
        print("   Reading files into memory as full DataFrames.\n")

    # Read all files and track failures
    results = [(f, read_file(f, chunk_size=chunk_size, verbose=False)) for f in tqdm(files, desc="Merging files")]

    # Separate successes and failures
    failed_files = [(f.name, err) for f, (df, err) in results if err is not None]
    successful_dfs = [df for f, (df, err) in results if err is None]

    merged_df = pd.concat(successful_dfs, ignore_index=True)

    print(f"\n📐 Merged DataFrame shape: {merged_df.shape}")
    print(f"✅ Files merged: {len(successful_dfs)} of {len(files)}")

    # Report failures explicitly
    if failed_files:
        print(f"\n⚠️  FAILED FILES ({len(failed_files)}):")
        for fname, err in failed_files:
            print(f"   • {fname}: {err}")
    print()

    if show_preview:
        missing_counts = merged_df.isnull().sum()
        missing_pct = (missing_counts / len(merged_df)) * 100
        top_missing = (
            pd.DataFrame(
                {
                    "Absent Count": missing_counts,
                    "Absent Percentage": missing_pct,
                }
            )
            .sort_values("Absent Count", ascending=False)
            .head(10)
        )

        print("🔟 TOP 10 COLUMNS WITH MISSING VALUES\n")
        with pd.option_context("display.max_rows", None):
            print(top_missing.to_string())
        print()

    if save:
        timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        output_file = DATASETS_OUTPUT / f"merged_{folder.name}_{timestamp}.csv"
        merged_df.to_csv(output_file, index=False)
        print(f"💾 Merged CSV saved to: {output_file}\n")

    return merged_df

# ======================================================
# 🏁 Script execution
# ======================================================

if __name__ == "__main__":
    print("🧪 Running Phase 1 preview...")
    datasets_overview()
    print("\nNext steps (uncomment to run):")
    print("   # convert_excel_to_csv()")
    print("   # df = merge_folder_files(save=True)")
    print("\n💡 Tip: Run 'python dataset_ingestion.py' for Phase 1 demo")
